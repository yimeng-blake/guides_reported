#!/usr/bin/env python3
"""
Earnings Guidance Philosophy Analyzer
Comprehensive analysis of management's guidance track record and philosophy.

Data source: financialdatasets.ai (8-K filings, income statements, stock prices)
Parser: Claude API (LLM-based, format-agnostic)

Output: Multi-sheet Excel workbook with:
  1. Executive Dashboard
  2. Guidance vs Actuals (quarterly detail)
  3. Full-Year Guidance Walk
  4. Conservatism Score & Trend
  5. Stock Price Reaction
  6. Seasonal Patterns
  7. Multi-Metric Accuracy
"""

import sys
import json
import hashlib
import math
import os
import re
import time
import threading
import requests
from concurrent.futures import ThreadPoolExecutor, as_completed
from datetime import datetime, timedelta
from html.parser import HTMLParser
from pathlib import Path

import anthropic
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, Reference, LineChart, ScatterChart
from openpyxl.chart.series import DataPoint

# ── Config ──────────────────────────────────────────────────────────────
# Read API keys from Streamlit secrets (cloud) or fall back to env vars (local dev)
def _get_secret(key: str) -> str:
    """Read from st.secrets → env var. Raises if neither is set."""
    # Try Streamlit secrets first
    try:
        import streamlit as st
        if hasattr(st, "secrets") and key in st.secrets:
            return st.secrets[key]
    except Exception:
        pass
    # Fall back to environment variable
    val = os.environ.get(key, "")
    if not val:
        raise RuntimeError(
            f"Missing required secret: {key}. "
            f"Set it in .streamlit/secrets.toml or as an environment variable."
        )
    return val

FD_API_KEY = _get_secret("FD_API_KEY")
CLAUDE_API_KEY = _get_secret("ANTHROPIC_API_KEY")
BASE_URL = "https://api.financialdatasets.ai"
HEADERS = {"X-API-KEY": FD_API_KEY}

claude_client = anthropic.Anthropic(api_key=CLAUDE_API_KEY)


# ── Adaptive Rate Limiter ──────────────────────────────────────────────
class RateLimiter:
    """Thread-safe adaptive rate limiter. Shared across all API call threads."""
    def __init__(self, min_interval: float = 0.6, max_interval: float = 5.0):
        self._lock = threading.Lock()
        self._interval = min_interval
        self._min = min_interval
        self._max = max_interval
        self._last_request = 0.0

    def wait(self):
        """Block until it's safe to make the next request."""
        with self._lock:
            now = time.monotonic()
            elapsed = now - self._last_request
            if elapsed < self._interval:
                time.sleep(self._interval - elapsed)
            self._last_request = time.monotonic()

    def back_off(self):
        """Widen the interval after a 429 or transient error."""
        with self._lock:
            self._interval = min(self._interval * 1.5, self._max)

    def ease_up(self):
        """Gradually reduce interval after a successful request."""
        with self._lock:
            self._interval = max(self._interval * 0.9, self._min)

    @property
    def current_interval(self):
        return self._interval

_fd_rate_limiter = RateLimiter(min_interval=1.0, max_interval=8.0)

# ── Styles ──────────────────────────────────────────────────────────────
BEAT_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
MISS_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
RANGE_FILL = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
NO_GUIDE_FILL = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
EST_FILL = PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid")
HEADER_FILL = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
SECTION_FILL = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=10)
TITLE_FONT = Font(bold=True, size=14, color="1F4E79")
SECTION_FONT = Font(bold=True, size=12, color="1F4E79")
LABEL_FONT = Font(bold=True, size=11)
THIN_BORDER = Border(
    left=Side(style="thin", color="D9D9D9"),
    right=Side(style="thin", color="D9D9D9"),
    top=Side(style="thin", color="D9D9D9"),
    bottom=Side(style="thin", color="D9D9D9"),
)
VERDICT_FONT = {
    "BEAT": Font(bold=True, color="006100"),
    "MISS": Font(bold=True, color="9C0006"),
    "IN-RANGE": Font(bold=True, color="9C6500"),
}
VERDICT_FILL = {"BEAT": BEAT_FILL, "MISS": MISS_FILL, "IN-RANGE": RANGE_FILL, "NO GUIDANCE": NO_GUIDE_FILL}
RAISE_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
LOWER_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
REAFFIRM_FILL = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")


# ── LLM parse cache (dual-layer: filesystem + session state) ──────────
CACHE_DIR = Path.home() / ".cache" / "guidance_analyzer"


def _session_cache() -> dict | None:
    """Return the st.session_state cache dict, or None if not in Streamlit."""
    try:
        import streamlit as st
        if "guidance_cache" not in st.session_state:
            st.session_state["guidance_cache"] = {}
        return st.session_state["guidance_cache"]
    except Exception:
        return None


def _cache_key(ticker: str, filing_date: str, text: str) -> Path:
    """Build a cache file path from ticker, date, and a content hash."""
    h = hashlib.md5(text.encode("utf-8", errors="replace")).hexdigest()[:12]
    return CACHE_DIR / f"{ticker}_{filing_date}_{h}.json"


def _cache_get(ticker: str, filing_date: str, text: str | None = None) -> dict | None:
    """Return cached LLM parse result if it exists.

    Checks filesystem first, then falls back to Streamlit session state.
    If text is provided, uses exact content-hash match for filesystem.
    If text is None, looks up by ticker+date prefix (any hash).
    """
    # Layer 1: Filesystem cache
    if text is not None:
        p = _cache_key(ticker, filing_date, text)
        if p.exists():
            try:
                return json.loads(p.read_text())
            except (json.JSONDecodeError, OSError):
                pass
    # Filesystem fallback: find any cache file for this ticker+date
    try:
        prefix = f"{ticker}_{filing_date}_"
        for p in CACHE_DIR.iterdir():
            if p.name.startswith(prefix) and p.suffix == ".json":
                try:
                    return json.loads(p.read_text())
                except (json.JSONDecodeError, OSError):
                    continue
    except (OSError, FileNotFoundError):
        pass

    # Layer 2: Streamlit session state (survives re-runs on Cloud)
    sc = _session_cache()
    if sc is not None:
        session_key = f"{ticker}_{filing_date}"
        if session_key in sc:
            return sc[session_key]

    return None


def _cache_put(ticker: str, filing_date: str, text: str, result: dict):
    """Write LLM parse result to both filesystem and session state."""
    # Layer 1: Filesystem
    try:
        CACHE_DIR.mkdir(parents=True, exist_ok=True)
        p = _cache_key(ticker, filing_date, text)
        p.write_text(json.dumps(result))
    except OSError:
        pass  # Non-fatal

    # Layer 2: Session state
    sc = _session_cache()
    if sc is not None:
        session_key = f"{ticker}_{filing_date}"
        sc[session_key] = result


# ── API helpers ─────────────────────────────────────────────────────────

def fetch_8k_filings(ticker: str, limit: int = 200) -> list[dict]:
    r = requests.get(f"{BASE_URL}/filings",
        params={"ticker": ticker, "filing_type": "8-K", "limit": limit},
        headers=HEADERS, timeout=30)
    if r.status_code in (401, 403):
        raise RuntimeError(f"FD API authentication failed (HTTP {r.status_code}). Check your FD_API_KEY.")
    r.raise_for_status()
    return r.json().get("filings", [])

_RETRYABLE_STATUS = {429, 500, 502, 503, 504}

def fetch_exhibit_text(ticker: str, accession: str, retries: int = 5) -> str | None:
    for attempt in range(retries + 1):
        _fd_rate_limiter.wait()  # Global pacing — prevents 429 storms
        try:
            r = requests.get(f"{BASE_URL}/filings/items",
                params={"ticker": ticker, "filing_type": "8-K",
                        "accession_number": accession, "include_exhibits": True},
                headers=HEADERS, timeout=30)
        except (requests.exceptions.Timeout, requests.exceptions.ConnectionError):
            _fd_rate_limiter.back_off()
            if attempt < retries:
                time.sleep(2 * (attempt + 1))
                continue
            return None
        if r.status_code == 200:
            _fd_rate_limiter.ease_up()
            try:
                for item in r.json().get("items", []):
                    for ex in item.get("exhibits", []):
                        text = ex.get("text", "")
                        if len(text) > 3000:
                            return text
            except (ValueError, KeyError):
                pass  # Bad JSON
            if attempt < retries:
                time.sleep(1.5)
                continue
            return None
        if r.status_code in _RETRYABLE_STATUS:
            _fd_rate_limiter.back_off()
            if attempt < retries:
                time.sleep(2 * (attempt + 1))
                continue
        return None
    return None


# ── SEC EDGAR fallback for older filings ──────────────────────────────

EDGAR_HEADERS = {"User-Agent": "GuidanceAnalyzer research@example.com"}

class _HTMLTextExtractor(HTMLParser):
    """Strip HTML tags, return plain text."""
    def __init__(self):
        super().__init__()
        self._parts: list[str] = []
        self._skip = False
    def handle_starttag(self, tag, attrs):
        if tag in ("script", "style", "ix:header"):
            self._skip = True
    def handle_endtag(self, tag):
        if tag in ("script", "style", "ix:header"):
            self._skip = False
        if tag in ("p", "div", "tr", "br", "li", "h1", "h2", "h3", "h4"):
            self._parts.append("\n")
    def handle_data(self, data):
        if not self._skip:
            self._parts.append(data)
    def get_text(self) -> str:
        raw = "".join(self._parts)
        # Collapse whitespace runs but keep newlines
        raw = re.sub(r"[^\S\n]+", " ", raw)
        raw = re.sub(r"\n{3,}", "\n\n", raw)
        return raw.strip()

def _html_to_text(html: str) -> str:
    parser = _HTMLTextExtractor()
    parser.feed(html)
    return parser.get_text()


def _edgar_get(url: str, timeout: int = 15, retries: int = 2) -> requests.Response | None:
    """GET with simple retry for EDGAR. Returns None on total failure."""
    for attempt in range(retries + 1):
        try:
            r = requests.get(url, headers=EDGAR_HEADERS, timeout=timeout)
            if r.status_code == 200:
                return r
            if r.status_code in (429, 500, 502, 503) and attempt < retries:
                time.sleep(1.5 * (attempt + 1))
                continue
            return r  # Non-retryable status
        except (requests.exceptions.Timeout, requests.exceptions.ConnectionError):
            if attempt < retries:
                time.sleep(1.5 * (attempt + 1))
                continue
    return None


def _get_cik_for_ticker(ticker: str) -> str | None:
    """Resolve a stock ticker to an SEC CIK number."""
    try:
        r = _edgar_get("https://www.sec.gov/files/company_tickers.json", timeout=10)
        if not r or r.status_code != 200:
            return None
        for entry in r.json().values():
            if entry.get("ticker", "").upper() == ticker.upper():
                return str(entry["cik_str"]).zfill(10)
    except Exception:
        pass
    return None


def _fetch_edgar_8k_filings(ticker: str, cik: str) -> list[dict]:
    """Fetch ALL 8-K filing metadata from SEC EDGAR (goes back to IPO)."""
    all_filings = []
    try:
        # Recent filings
        r = _edgar_get(f"https://data.sec.gov/submissions/CIK{cik}.json")
        if not r or r.status_code != 200:
            return all_filings
        data = r.json()

        def _extract_8ks(rec: dict) -> list[dict]:
            results = []
            forms = rec.get("form", [])
            dates = rec.get("filingDate", [])
            accessions = rec.get("accessionNumber", [])
            primary_docs = rec.get("primaryDocument", [""] * len(forms))
            for i in range(len(forms)):
                if forms[i] == "8-K":
                    results.append({
                        "filing_date": dates[i],
                        "accession_number": accessions[i],
                        "primary_doc": primary_docs[i] if i < len(primary_docs) else "",
                        "cik": cik,
                    })
            return results

        all_filings.extend(_extract_8ks(data["filings"]["recent"]))

        # Older filings in supplemental files
        for f in data["filings"].get("files", []):
            fname = f.get("name", "")
            if not fname:
                continue
            r2 = _edgar_get(f"https://data.sec.gov/submissions/{fname}")
            if r2 and r2.status_code == 200:
                all_filings.extend(_extract_8ks(r2.json()))
            time.sleep(0.2)  # Be polite to EDGAR

    except Exception as e:
        print(f"  EDGAR fetch error: {e}")

    all_filings.sort(key=lambda x: x["filing_date"])
    return all_filings


def _fetch_edgar_exhibit_text(filing: dict) -> str | None:
    """Fetch the earnings press release exhibit from an EDGAR 8-K filing.

    The exhibit (earnings PR) is typically the largest .htm file in the filing,
    distinct from the 8-K form itself (primary_doc).
    """
    cik = filing["cik"].lstrip("0")
    acc_path = filing["accession_number"].replace("-", "")

    try:
        # Get the filing index to find all documents
        idx_url = f"https://www.sec.gov/Archives/edgar/data/{cik}/{acc_path}/index.json"
        r = _edgar_get(idx_url)
        if not r or r.status_code != 200:
            return None

        items = r.json().get("directory", {}).get("item", [])
        # Find .htm files, excluding index files
        htm_files = []
        for item in items:
            name = item.get("name", "")
            size_str = item.get("size", "0")
            try:
                size = int(size_str) if size_str else 0
            except (ValueError, TypeError):
                size = 0
            if name.endswith(".htm") and "index" not in name:
                htm_files.append((name, size))

        if not htm_files:
            return None

        # The earnings exhibit is the largest .htm — it contains the full PR
        htm_files.sort(key=lambda x: -x[1])
        exhibit_name = htm_files[0][0]

        url = f"https://www.sec.gov/Archives/edgar/data/{cik}/{acc_path}/{exhibit_name}"
        r = _edgar_get(url, timeout=20)
        if not r or r.status_code != 200:
            return None

        text = _html_to_text(r.text)
        if len(text) > 3000:
            return text
    except Exception as e:
        print(f"    EDGAR exhibit fetch error: {e}")
    return None


def fetch_income_statements(ticker: str, limit: int = 30) -> list[dict]:
    r = requests.get(f"{BASE_URL}/financials/income-statements",
        params={"ticker": ticker, "period": "quarterly", "limit": limit},
        headers=HEADERS)
    r.raise_for_status()
    return r.json().get("income_statements", [])


def fetch_stock_prices(ticker: str, start_date: str, end_date: str) -> list[dict]:
    try:
        r = requests.get(f"{BASE_URL}/prices",
            params={"ticker": ticker, "interval": "day",
                    "start_date": start_date, "end_date": end_date},
            headers=HEADERS, timeout=15)
        if r.status_code != 200:
            return []
        return r.json().get("prices", [])
    except (requests.exceptions.Timeout, requests.exceptions.ConnectionError, ValueError):
        return []


# ── LLM Parser ─────────────────────────────────────────────────────────

LLM_PARSE_PROMPT = """You are a financial data extraction assistant. Given an earnings press release (8-K exhibit), extract the following structured data. Return ONLY valid JSON, no markdown, no commentary.

INSTRUCTIONS:
1. Identify which fiscal quarter this press release reports results for.
2. Extract the ACTUAL revenue reported for that quarter (in millions USD). Use the company's PRIMARY guided revenue metric — this varies by company:
   - Snowflake: "Product revenue"
   - ServiceNow: "Subscription revenues"
   - Walmart: "Net sales"
   - Most others: "Total revenue" or just "Revenue"
   The key is to use the SAME metric that the company gives guidance on, so we can compare apples to apples.
3. Extract NEXT-QUARTER guidance for revenue (the quarter AFTER the one being reported). This is forward-looking guidance in the "Guidance", "Outlook", "Business Outlook", or "Financial Outlook" section. Revenue guidance is typically a dollar range like "$X million to $Y million" or "$X billion to $Y billion" or "$X - $Y". Convert everything to millions USD. Use the same metric from step 2 (e.g., if the company guides on subscription revenue, extract subscription revenue guidance).
4. Extract FULL-YEAR guidance for the relevant fiscal year's revenue. This may be a dollar amount range or a growth rate range. Convert dollar amounts to millions USD. Use the same revenue metric as steps 2-3.
5. Extract any operating margin guidance (as a percentage). Companies typically guide on NON-GAAP operating margin.
6. Extract the ACTUAL non-GAAP operating margin for the reported quarter (as a percentage). This is usually found in the non-GAAP reconciliation tables or summary financial highlights. Look for "non-GAAP income from operations" divided by revenue, or an explicit "non-GAAP operating margin" line. If only GAAP operating margin is available, use that instead.
7. Determine if the company uses fiscal years (e.g., "Fiscal 2026" ending Jan 31) or calendar years.

IMPORTANT:
- "Product revenue" guidance and actuals should be used for Snowflake. For most other companies, use total revenue.
- Guidance is FORWARD-LOOKING: if reporting Q3 results, the next-quarter guidance is for Q4.
- Full-year guidance may reference the current fiscal year (being updated) or the next fiscal year (if reporting Q4 and giving initial FY guidance).
- EVERY earnings release with a "Guidance" or "Outlook" section has full-year guidance — look carefully. Q4 earnings give the NEXT fiscal year's initial FY guide. Q1/Q2/Q3 earnings update the CURRENT fiscal year's FY guide.
- If guidance is a SINGLE number (not a range), use that number for BOTH the low and high fields. For example, if FY guidance says "Product revenue $5,660" with no range, set both fy_rev_guide_low_millions and fy_rev_guide_high_millions to 5660.0.
- If guidance is given as a growth rate (e.g., "net sales expected to grow 3% to 4%"), report it in the growth fields, not the dollar fields.
- Revenue numbers in financial statement tables may be in thousands — check the table header for "(in thousands)" or "(in millions)" and convert to millions.
- If you truly cannot find a field after careful inspection, use null.

CRITICAL — Full-year vs quarterly confusion:
- The fy_rev_guide fields MUST contain FULL-YEAR (annual) revenue guidance, NOT quarterly revenue.
- Full-year revenue guidance should be roughly 4x the quarterly revenue. If you find a number that looks like a single quarter's revenue, do NOT put it in the FY fields.
- Use the SAME revenue metric consistently for both quarterly actuals AND full-year guidance. If the company guides on "subscription revenues", use subscription revenues for BOTH. Do NOT mix subscription revenue actuals with total revenue guidance or vice versa.
- ServiceNow (NOW) guides on "Subscription revenues" — use subscription revenues for all fields.

CRITICAL — CY vs FY prefix:
- Use "FY" ONLY if the company explicitly uses "Fiscal Year" or "Fiscal" in their reporting (e.g., Snowflake says "Fiscal 2026", ServiceNow says "Fiscal Year 2025"). These companies have fiscal years ending in January/February.
- Use "CY" for ALL companies that report using calendar years (January-December), even if their Q4 report title says "Full-Year 2025 Results". A Q4 report covering "year ended December 31" is CY, not FY.
- Be CONSISTENT across all filings for the same company. If Q1/Q2/Q3 are CY, then Q4 must also be CY.

Return this exact JSON structure:
{
  "reported_quarter": "FY2026-Q4" or "CY2025-Q3" (use FY for fiscal year companies, CY for calendar year),
  "actual_revenue_millions": 695.1,
  "actual_non_gaap_op_margin_pct": 9.5 or null,
  "revenue_metric_name": "Total revenue" or "Product revenue" or "Net sales",
  "next_q_target": "FY2027-Q1" or "CY2026-Q1",
  "next_q_rev_guide_low_millions": 659.0,
  "next_q_rev_guide_high_millions": 664.0,
  "next_q_op_margin_guide_pct": 9.0 or null,
  "fy_target": "FY2027" or "CY2026",
  "fy_rev_guide_low_millions": 2860.0 or null,
  "fy_rev_guide_high_millions": 2900.0 or null,
  "fy_rev_growth_low_pct": null or 3.0,
  "fy_rev_growth_high_pct": null or 4.0,
  "fy_op_margin_guide_pct": 12.5 or null,
  "fy_fcf_margin_guide_pct": 23.0 or null,
  "fy_eps_guide_low": 2.75 or null,
  "fy_eps_guide_high": 2.85 or null,
  "is_earnings_release": true
}
"""


def llm_parse_filing(text: str, ticker: str, log_fn=None) -> dict | None:
    """Use Claude to extract structured data from an 8-K exhibit."""
    # Smart truncation: always include the guidance/outlook section even if it's
    # far into the document. Many companies put guidance after the financial tables.
    if len(text) > 25000:
        lower = text.lower()
        # Find the guidance section start
        guidance_start = -1
        for kw in ["financial outlook", "business outlook", "guidance\nbased on",
                    "guidance\nthe following", "guidance\nour guidance",
                    "first quarter", "second quarter", "third quarter", "fourth quarter"]:
            idx = lower.find(kw, 5000)  # skip the header area
            if idx != -1 and "guidance" in lower[max(0,idx-200):idx+500]:
                guidance_start = idx
                break
        if guidance_start == -1:
            # Fallback: look for any "guidance" near a dollar range
            for kw in ["guidance", "outlook"]:
                idx = lower.find(kw, 5000)
                if idx != -1:
                    nearby = text[idx:idx+1000]
                    if "$" in nearby and ("million" in nearby.lower() or "billion" in nearby.lower() or "%" in nearby):
                        guidance_start = idx
                        break

        if guidance_start > 14000:
            # Guidance is deep in the filing — keep: header + guidance section + financial tables at end
            text = (text[:8000]
                    + "\n\n[...truncated...]\n\n"
                    + text[max(8000, guidance_start - 500):guidance_start + 4000]
                    + "\n\n[...truncated...]\n\n"
                    + text[-8000:])
        else:
            # Guidance is in the first 14k — standard truncation
            text = text[:16000] + "\n\n[...middle section truncated...]\n\n" + text[-8000:]

    for attempt in range(4):
        try:
            resp = claude_client.messages.create(
                model="claude-sonnet-4-20250514",
                max_tokens=1000,
                messages=[{
                    "role": "user",
                    "content": f"Ticker: {ticker}\n\nPress release text:\n{text}"
                }],
                system=LLM_PARSE_PROMPT,
            )
            raw = resp.content[0].text.strip()
            # Strip markdown code fences if present
            if raw.startswith("```"):
                raw = raw.split("\n", 1)[1]
                if raw.endswith("```"):
                    raw = raw[:-3]
                elif "```" in raw:
                    raw = raw[:raw.rfind("```")]
            return json.loads(raw)
        except anthropic.RateLimitError:
            wait = 15 * (attempt + 1)
            msg = f"    Rate limited, retrying in {wait}s (attempt {attempt+1}/4)..."
            print(msg)
            if log_fn:
                log_fn(msg)
            time.sleep(wait)
        except anthropic.APIStatusError as e:
            # Retry on 500/502/503/529 (transient server errors)
            if e.status_code in (500, 502, 503, 529) and attempt < 3:
                wait = 10 * (attempt + 1)
                print(f"    API error {e.status_code}, retrying in {wait}s (attempt {attempt+1}/4)...")
                time.sleep(wait)
            else:
                print(f"    LLM API error (non-retryable): {e}")
                return None
        except Exception as e:
            print(f"    LLM parse error: {e}")
            return None
    print(f"    Failed after 4 retries")
    return None


def is_earnings_8k_quick(text: str) -> bool:
    """Fast heuristic check: is this an earnings press release?
    Uses a scoring system: need at least 2 of 3 signals to qualify."""
    lower = text[:8000].lower()
    has_quarter = any(q in lower for q in [
        "first quarter", "second quarter", "third quarter", "fourth quarter",
        " q1 ", " q2 ", " q3 ", " q4 ",
        "q1 ", "q2 ", "q3 ", "q4 ",  # start of line
        "three months ended", "six months ended", "twelve months ended",
        "full year", "full-year", "fiscal year",
    ])
    has_results = any(kw in lower for kw in [
        "results", "financial results", "announces", "reported",
        "results of operations", "quarterly report", "earnings",
    ])
    has_revenue = any(kw in lower for kw in [
        "revenue", "net sales", "product revenue", "subscription revenue",
        "total revenue", "net revenue", "billings",
    ])
    # Need at least 2 of 3 signals (some filings omit "revenue" from the header)
    score = sum([has_quarter, has_results, has_revenue])
    return score >= 2


# ── Math helpers ────────────────────────────────────────────────────────

def linear_regression(xs, ys):
    n = len(xs)
    if n < 2:
        return 0, 0
    sx = sum(xs)
    sy = sum(ys)
    sxy = sum(x * y for x, y in zip(xs, ys))
    sx2 = sum(x * x for x in xs)
    denom = n * sx2 - sx * sx
    if denom == 0:
        return 0, sy / n
    slope = (n * sxy - sx * sy) / denom
    intercept = (sy - slope * sx) / n
    return slope, intercept


def correlation(xs, ys):
    n = len(xs)
    if n < 3:
        return None
    mx = sum(xs) / n
    my = sum(ys) / n
    num = sum((x - mx) * (y - my) for x, y in zip(xs, ys))
    dx = math.sqrt(sum((x - mx) ** 2 for x in xs))
    dy = math.sqrt(sum((y - my) ** 2 for y in ys))
    if dx == 0 or dy == 0:
        return None
    return num / (dx * dy)


# ── Core analysis ───────────────────────────────────────────────────────

def build_all_data(ticker: str, progress_callback=None) -> dict:
    """Master function that fetches and parses everything."""
    def log(msg):
        print(msg)
        if progress_callback:
            progress_callback(msg)

    log("Fetching 8-K filings from financialdatasets.ai...")
    filings = fetch_8k_filings(ticker)

    # Determine the earliest date covered by the primary API
    fd_dates = sorted(f.get("report_date", "") for f in filings if f.get("report_date"))
    fd_earliest = fd_dates[0] if fd_dates else "9999-01-01"

    # Supplement with SEC EDGAR for older filings
    edgar_filings = []
    cik = _get_cik_for_ticker(ticker)
    if cik:
        log(f"Fetching older 8-K filings from SEC EDGAR (CIK {cik})...")
        all_edgar = _fetch_edgar_8k_filings(ticker, cik)
        # Only keep EDGAR filings older than what the primary API has
        edgar_filings = [f for f in all_edgar if f["filing_date"] < fd_earliest]
        if edgar_filings:
            log(f"  Found {len(edgar_filings)} additional 8-K filings from EDGAR ({edgar_filings[0]['filing_date']} to {edgar_filings[-1]['filing_date']})")
        else:
            log("  No additional older filings found on EDGAR.")
    else:
        log("  Could not resolve CIK for EDGAR lookup.")

    log("Fetching income statements...")
    stmts = fetch_income_statements(ticker, limit=80)
    stmt_by_fp = {}
    for s in stmts:
        fp = s.get("fiscal_period", "")
        if fp and not fp.startswith("FY") and not fp.startswith("CY"):
            fp = "FY" + fp
        stmt_by_fp[fp] = s

    # ── Phase 1: Fetch all exhibit texts ──────────────────────────────
    # Each item: (filing_date, text)
    filing_texts = []

    # EDGAR older filings (sequential — SEC rate limit)
    if edgar_filings:
        # Check cache first — already-parsed filings don't need re-fetching
        cached_dates = set()
        for ef in edgar_filings:
            fd = ef.get("filing_date", "")
            if _cache_get(ticker, fd):
                cached_dates.add(fd)

        log(f"Fetching {len(edgar_filings)} older EDGAR 8-K exhibit texts...")

        for i, filing in enumerate(edgar_filings):
            fd = filing.get("filing_date", "?")
            log(f"  Fetching EDGAR filing {i+1}/{len(edgar_filings)}...")

            # Skip if already cached (just need to recover from cache later)
            if fd in cached_dates:
                continue

            try:
                text = _fetch_edgar_exhibit_text(filing)
            except Exception as e:
                print(f"  [warn] EDGAR fetch failed for {fd}: {e}")
                text = None
            if text and is_earnings_8k_quick(text):
                filing_texts.append((filing["filing_date"], text))
            elif text:
                log(f"  [skipped] {fd} — failed earnings heuristic")

    # Primary API filings (concurrent with modest parallelism to avoid 429s)
    total_primary = len(filings)
    log(f"Fetching {total_primary} 8-K exhibit texts from primary API...")

    _FETCH_SKIPPED = "SKIPPED"  # Sentinel: text fetched but not earnings

    def _fetch_primary_text(filing):
        acc = filing["accession_number"]
        fd = filing.get("report_date", "?")
        try:
            text = fetch_exhibit_text(ticker, acc)
        except Exception:
            return None  # Genuine fetch failure
        if text and is_earnings_8k_quick(text):
            return (fd, text)
        if text:
            print(f"  [skipped] {fd} — failed earnings heuristic")
            return _FETCH_SKIPPED  # Not earnings — don't retry
        return None  # No text returned — fetch failed

    fetch_done = [0]
    failed_filings = []
    with ThreadPoolExecutor(max_workers=2) as pool:
        futures = {pool.submit(_fetch_primary_text, f): f for f in filings}
        for future in as_completed(futures):
            fetch_done[0] += 1
            if fetch_done[0] % 10 == 0 or fetch_done[0] == total_primary:
                log(f"  Fetched {fetch_done[0]}/{total_primary} exhibit texts...")
            try:
                result = future.result()
                if result and result != _FETCH_SKIPPED:
                    filing_texts.append(result)
                elif result is None:
                    # Genuine fetch failure — worth retrying
                    failed_filings.append(futures[future])
                # _FETCH_SKIPPED means not earnings — don't retry
            except Exception as e:
                filing = futures[future]
                failed_filings.append(filing)
                print(f"  [warn] Failed to fetch {filing.get('report_date', '?')}: {e}")

    # ── Second pass: retry failed fetches sequentially with generous pacing ──
    if failed_filings:
        log(f"Retrying {len(failed_filings)} failed exhibit fetches (sequential)...")
        _fd_rate_limiter.back_off()  # Preemptively slow down
        for i, filing in enumerate(failed_filings):
            time.sleep(2.0)  # Generous gap between retries
            acc = filing.get("accession_number", "")
            fd = filing.get("report_date", "?")
            try:
                text = fetch_exhibit_text(ticker, acc, retries=3)
                if text and is_earnings_8k_quick(text):
                    filing_texts.append((fd, text))
                    log(f"  [recovered] {fd} ({i+1}/{len(failed_filings)})")
                elif text:
                    pass  # Not an earnings release, fine to skip
            except Exception:
                pass  # Already logged in first pass

    log(f"Found {len(filing_texts)} earnings releases to parse.")

    # ── Phase 2: Parse with LLM (cached hits instant, misses concurrent) ──
    def _result_to_entry(filing_date, result):
        """Convert raw LLM result dict to a parsed entry."""
        return {
            "filing_date": filing_date,
            "reported_q": result.get("reported_quarter"),
            "actual_revenue": result.get("actual_revenue_millions"),
            "actual_non_gaap_op_margin": result.get("actual_non_gaap_op_margin_pct"),
            "revenue_metric": result.get("revenue_metric_name", "Revenue"),
            "guide_target_q": result.get("next_q_target"),
            "guide_low": result.get("next_q_rev_guide_low_millions"),
            "guide_high": result.get("next_q_rev_guide_high_millions"),
            "guide_op_margin": result.get("next_q_op_margin_guide_pct"),
            "fy_target": result.get("fy_target"),
            "fy_rev_low": result.get("fy_rev_guide_low_millions"),
            "fy_rev_high": result.get("fy_rev_guide_high_millions"),
            "fy_rev_growth_low": result.get("fy_rev_growth_low_pct"),
            "fy_rev_growth_high": result.get("fy_rev_growth_high_pct"),
            "fy_op_margin": result.get("fy_op_margin_guide_pct"),
            "fy_fcf_margin": result.get("fy_fcf_margin_guide_pct"),
            "fy_eps_low": result.get("fy_eps_guide_low"),
            "fy_eps_high": result.get("fy_eps_guide_high"),
        }

    parsed = []
    to_parse = []  # (filing_date, text) items that need LLM parsing
    fetched_dates = {fd for fd, _ in filing_texts}

    # Check cache first — for texts we fetched, use exact match;
    # also recover any previously-cached dates whose text fetch failed this time
    cached_count = 0
    cached_dates = set()
    for filing_date, text in filing_texts:
        cached = _cache_get(ticker, filing_date, text)
        if cached is not None:
            if cached.get("is_earnings_release"):
                parsed.append(_result_to_entry(filing_date, cached))
            cached_count += 1
            cached_dates.add(filing_date)
            log(f"  [cached] {filing_date}")
        else:
            to_parse.append((filing_date, text))

    # Recover filings that were cached from a previous run but whose text
    # fetch failed this time (e.g. due to API rate limiting)
    all_known_dates = set()
    for f in edgar_filings:
        all_known_dates.add(f["filing_date"])
    for f in filings:
        if f.get("report_date"):
            all_known_dates.add(f["report_date"])
    missed_dates = all_known_dates - fetched_dates - cached_dates
    recovered = 0
    for fd in sorted(missed_dates):
        cached = _cache_get(ticker, fd)  # date-only lookup, no text
        if cached is not None and cached.get("is_earnings_release"):
            parsed.append(_result_to_entry(fd, cached))
            recovered += 1
            log(f"  [recovered from cache] {fd}")
    if recovered:
        log(f"  Recovered {recovered} filings from cache (text fetch had failed).")

    if cached_count:
        log(f"  Loaded {cached_count} filings from cache.")

    if to_parse:
        log(f"Parsing {len(to_parse)} filings with Claude AI (concurrent)...")
        parse_done = [0]

        def _parse_one(item):
            filing_date, text = item
            # No log() calls here — Streamlit widgets can't be updated from worker threads
            result = llm_parse_filing(text, ticker)
            if result:
                _cache_put(ticker, filing_date, text, result)
            return filing_date, result

        with ThreadPoolExecutor(max_workers=4) as pool:
            futures = {pool.submit(_parse_one, item): item for item in to_parse}
            for future in as_completed(futures):
                parse_done[0] += 1
                try:
                    filing_date, result = future.result()
                    log(f"  Parsing {filing_date}... ({parse_done[0]}/{len(to_parse)})")
                    if result and result.get("is_earnings_release"):
                        parsed.append(_result_to_entry(filing_date, result))
                except Exception as e:
                    item = futures[future]
                    log(f"  [warn] Parse failed for {item[0]}: {e} ({parse_done[0]}/{len(to_parse)})")

    log("Normalizing and validating parsed data...")
    parsed.sort(key=lambda x: x["filing_date"])

    # ── Normalize CY/FY prefixes ────────────────────────────────────
    # The LLM sometimes inconsistently uses CY vs FY across filings for the
    # same company (e.g., CY for Q1-Q3 but FY for Q4). Detect the majority
    # prefix and normalize everything to it.
    prefix_counts = {"CY": 0, "FY": 0}
    for p in parsed:
        rq = p.get("reported_q") or ""
        if rq.startswith("CY"): prefix_counts["CY"] += 1
        elif rq.startswith("FY"): prefix_counts["FY"] += 1

    if prefix_counts["CY"] > 0 and prefix_counts["FY"] > 0:
        # Mixed prefixes — normalize to the majority
        majority = "CY" if prefix_counts["CY"] >= prefix_counts["FY"] else "FY"
        minority = "FY" if majority == "CY" else "CY"
        log(f"  Normalizing mixed {minority}/{majority} prefixes → all {majority}")
        for p in parsed:
            for key in ["reported_q", "guide_target_q", "fy_target"]:
                val = p.get(key)
                if val and val.startswith(minority):
                    p[key] = majority + val[2:]

    # Log parsed entries (after normalization so prefixes are consistent)
    for entry in parsed:
        rev_str = f"${entry['actual_revenue']:,.0f}M" if entry["actual_revenue"] else "N/A"
        fy_label = entry["fy_target"] or "FY"
        fy_str = ""
        if entry["fy_rev_low"] is not None and entry["fy_rev_high"] is not None:
            mid = (entry["fy_rev_low"] + entry["fy_rev_high"]) / 2
            fy_str = f"{fy_label} guide: ${mid:,.0f}M"
        elif entry["fy_rev_low"] is not None:
            fy_str = f"{fy_label} guide: ${entry['fy_rev_low']:,.0f}M"
        elif entry["fy_rev_high"] is not None:
            fy_str = f"{fy_label} guide: ${entry['fy_rev_high']:,.0f}M"
        elif entry.get("fy_rev_growth_low") is not None:
            fy_str = f"{fy_label} growth: {entry['fy_rev_growth_low']}-{entry['fy_rev_growth_high']}%"
        else:
            fy_str = "no FY guide"
        log(f"    {entry['reported_q']}: {entry['revenue_metric']} {rev_str} | {fy_str}")

    # ── Build guidance lookup ────────────────────────────────────────
    guidance_by_target = {}
    for p in parsed:
        tq = p["guide_target_q"]
        if tq and p["guide_low"] is not None:
            guidance_by_target[tq] = {
                "guide_low": p["guide_low"],
                "guide_high": p["guide_high"],
                "guide_midpoint": (p["guide_low"] + p["guide_high"]) / 2,
                "guide_op_margin": p["guide_op_margin"],
                "guidance_given_in": p["reported_q"],
            }

    # Sanity-check quarterly guidance: if the guided midpoint is >2x the actual
    # revenue for that quarter, the LLM likely put a full-year number into the
    # quarterly guidance field. Drop these entries.
    for p in parsed:
        q = p["reported_q"]
        if not q:
            continue
        actual = p.get("actual_revenue")
        guide = guidance_by_target.get(q)
        if actual and guide and guide["guide_midpoint"] > actual * 2:
            log(f"  Dropping bogus quarterly guidance for {q}: "
                f"guided {guide['guide_midpoint']:.0f}M vs actual {actual:.0f}M "
                f"(likely FY number in quarterly field)")
            del guidance_by_target[q]

    # ── Module 0: Quarterly guidance vs actuals rows ─────────────────
    log("Building quarterly guidance vs actuals...")
    rows = []
    for p in parsed:
        q = p["reported_q"]
        if not q:
            continue
        actual_rev = p["actual_revenue"]
        guide = guidance_by_target.get(q)
        stmt = stmt_by_fp.get(q, {})
        total_rev_stmt = stmt.get("revenue")
        gross_profit = stmt.get("gross_profit")
        op_income = stmt.get("operating_income")
        net_income = stmt.get("net_income")
        total_rev = total_rev_stmt / 1e6 if total_rev_stmt and total_rev_stmt > 1e6 else total_rev_stmt
        gross_margin = (gross_profit / total_rev_stmt * 100) if total_rev_stmt and gross_profit else None
        # Prefer LLM-extracted non-GAAP op margin (matches how companies guide),
        # fall back to GAAP op margin from financial statements
        op_margin_actual = p.get("actual_non_gaap_op_margin")
        if op_margin_actual is None:
            op_margin_actual = (op_income / total_rev_stmt * 100) if total_rev_stmt and op_income else None

        if guide and actual_rev:
            midpoint = guide["guide_midpoint"]
            rev_diff = actual_rev - midpoint
            rev_diff_pct = (rev_diff / midpoint) * 100
            guide_low = guide["guide_low"]
            guide_high = guide["guide_high"]
            if actual_rev > guide_high:
                verdict = "BEAT"
            elif actual_rev < guide_low:
                verdict = "MISS"
            else:
                verdict = "IN-RANGE"
            guide_op_margin = guide["guide_op_margin"]
            guidance_source = guide["guidance_given_in"]
        else:
            midpoint = rev_diff = rev_diff_pct = guide_low = guide_high = None
            verdict = "NO GUIDANCE"
            guide_op_margin = guidance_source = None

        rows.append({
            "fiscal_quarter": q, "filing_date": p["filing_date"],
            "revenue_metric": p.get("revenue_metric", "Revenue"),
            "total_revenue": total_rev, "actual_revenue": actual_rev,
            "guide_low": guide_low, "guide_high": guide_high,
            "guide_midpoint": midpoint, "rev_diff_vs_mid": rev_diff,
            "rev_diff_pct": rev_diff_pct, "verdict": verdict,
            "guidance_source_q": guidance_source,
            "gross_profit": gross_profit, "gross_margin": gross_margin,
            "op_income": op_income, "op_margin_actual": op_margin_actual,
            "guide_op_margin": guide_op_margin, "net_income": net_income,
        })

    # YoY revenue growth
    for i, row in enumerate(rows):
        if i >= 4 and rows[i - 4]["actual_revenue"] and row["actual_revenue"]:
            row["rev_yoy"] = (row["actual_revenue"] / rows[i - 4]["actual_revenue"] - 1) * 100
        else:
            row["rev_yoy"] = None

    # ── Module 1: Full-year guidance walk ────────────────────────────
    log("Building full-year guidance walk...")
    fy_walk = {}
    for p in parsed:
        fy = p["fy_target"]
        if not fy:
            continue
        fy_rev = None
        if p["fy_rev_low"] is not None and p["fy_rev_high"] is not None:
            fy_rev = (p["fy_rev_low"] + p["fy_rev_high"]) / 2
        elif p["fy_rev_low"] is not None:
            # Single-point guidance (no range given)
            fy_rev = p["fy_rev_low"]
        elif p["fy_rev_high"] is not None:
            fy_rev = p["fy_rev_high"]
        if fy_rev is None:
            continue
        if fy not in fy_walk:
            fy_walk[fy] = []
        fy_walk[fy].append({
            "source_q": p["reported_q"],
            "filing_date": p["filing_date"],
            "fy_rev": fy_rev,
            "fy_rev_low": p["fy_rev_low"],
            "fy_rev_high": p["fy_rev_high"],
            "fy_op_margin": p.get("fy_op_margin"),
            "fy_fcf_margin": p.get("fy_fcf_margin"),
            "fy_eps_low": p.get("fy_eps_low"),
            "fy_eps_high": p.get("fy_eps_high"),
        })

    # Sanitize FY walk: multiple passes to catch different error types.

    # Pass 1: Ensure revisions are sorted by filing_date within each FY.
    for fy in fy_walk:
        fy_walk[fy].sort(key=lambda r: r["filing_date"])

    # Pass 2: Enforce that initial FY guidance must come from Q4 of prior year.
    # Rule: For CY20XX, initial guide is given in CY(XX-1)-Q4. Revisions are Q1-Q3 of XX.
    # Any entry not from this valid window is dropped.
    # If Q4 initial is missing (fetch failure), the FY still shows but starts from
    # the first available revision (Q1/Q2/Q3), clearly not the "initial".
    import re as _re
    for fy in list(fy_walk.keys()):
        m = _re.match(r"(CY|FY)(\d{4})", fy)
        if not m:
            continue
        prefix, year_str = m.group(1), int(m.group(2))
        revs = fy_walk[fy]

        # Filter to only valid source quarters
        cleaned = []
        for r in revs:
            sq = r["source_q"] or ""
            sq_m = _re.match(r"(CY|FY)(\d{4})-Q(\d)", sq)
            if sq_m:
                sq_year = int(sq_m.group(2))
                sq_q = int(sq_m.group(3))
                # Valid: Q4 of (year-1) = initial, Q1-Q3 of (year) = revisions
                is_valid = ((sq_year == year_str - 1 and sq_q == 4) or
                            (sq_year == year_str and sq_q in (1, 2, 3)))
                if is_valid:
                    cleaned.append(r)
                else:
                    log(f"  {fy}: dropped misattributed revision from {sq} "
                        f"(${r['fy_rev']:,.0f}M — wrong source quarter)")
            else:
                cleaned.append(r)

        if not cleaned:
            if revs:
                fy_walk[fy] = revs  # Don't empty
            continue
        fy_walk[fy] = cleaned

        # Enforce: first entry must be Q4 of prior year (the initial guide).
        # If it's not Q4, that means the Q4 filing was missed — don't show
        # a partial walk with a wrong "initial". Keep it but mark no initial.
        first_sq = cleaned[0].get("source_q", "")
        first_m = _re.match(r"(CY|FY)(\d{4})-Q(\d)", first_sq)
        if first_m:
            first_year = int(first_m.group(2))
            first_q = int(first_m.group(3))
            if not (first_year == year_str - 1 and first_q == 4):
                # No Q4 initial — this FY walk is incomplete.
                # Still keep it (shows revisions), but the dashboard
                # table will show it without a proper "initial" column.
                log(f"  {fy}: no Q4 initial guide found (walk starts from {first_sq})")

    # Pass 3: Drop outlier revisions that deviate >25% from median (metric mix-ups).
    for fy in list(fy_walk.keys()):
        revs = fy_walk[fy]
        if len(revs) < 2:
            continue
        vals = sorted(r["fy_rev"] for r in revs)
        median_val = vals[len(vals) // 2]
        cleaned = [r for r in revs if abs(r["fy_rev"] - median_val) / median_val < 0.25]
        if len(cleaned) < len(revs):
            dropped = len(revs) - len(cleaned)
            log(f"  {fy}: dropped {dropped} outlier revision(s) (likely LLM parse error)")
        if cleaned:
            fy_walk[fy] = cleaned
        else:
            # All revisions were outliers relative to each other — keep originals
            pass

    # Tag each FY with whether it has a proper Q4 initial
    _fy_has_q4_initial = {}
    for fy in fy_walk:
        m = _re.match(r"(CY|FY)(\d{4})", fy)
        if m and fy_walk[fy]:
            year_str = int(m.group(2))
            first_sq = fy_walk[fy][0].get("source_q", "")
            sq_m = _re.match(r"(CY|FY)(\d{4})-Q(\d)", first_sq)
            _fy_has_q4_initial[fy] = (sq_m and int(sq_m.group(2)) == year_str - 1
                                       and int(sq_m.group(3)) == 4)
        else:
            _fy_has_q4_initial[fy] = False

    fy_walk_rows = []
    for fy in sorted(fy_walk.keys()):
        revisions = fy_walk[fy]
        has_initial = _fy_has_q4_initial.get(fy, False)
        initial = revisions[0]["fy_rev"]
        for i, rev in enumerate(revisions):
            action = ("INITIAL" if i == 0 and has_initial else
                      "INITIAL" if i == 0 else  # No Q4 — still first entry
                      "RAISE" if rev["fy_rev"] > revisions[i-1]["fy_rev"] * 1.005 else
                      "LOWER" if rev["fy_rev"] < revisions[i-1]["fy_rev"] * 0.995 else
                      "REAFFIRM"
            )
            chg_vs_prior = rev["fy_rev"] - revisions[i-1]["fy_rev"] if i > 0 else None
            chg_vs_prior_pct = (chg_vs_prior / revisions[i-1]["fy_rev"] * 100) if chg_vs_prior and i > 0 else None
            chg_vs_initial = rev["fy_rev"] - initial if i > 0 else None
            chg_vs_initial_pct = (chg_vs_initial / initial * 100) if chg_vs_initial else None

            fy_walk_rows.append({
                "fy_target": fy, "revision_num": i,
                "source_q": rev["source_q"], "filing_date": rev["filing_date"],
                "fy_rev": rev["fy_rev"],
                "fy_rev_low": rev.get("fy_rev_low"),
                "fy_rev_high": rev.get("fy_rev_high"),
                "fy_op_margin": rev.get("fy_op_margin"),
                "fy_fcf_margin": rev.get("fy_fcf_margin"),
                "fy_eps_low": rev.get("fy_eps_low"),
                "fy_eps_high": rev.get("fy_eps_high"),
                "action": action,
                "chg_vs_prior": chg_vs_prior, "chg_vs_prior_pct": chg_vs_prior_pct,
                "chg_vs_initial": chg_vs_initial, "chg_vs_initial_pct": chg_vs_initial_pct,
            })

    # ── Module 2: Conservatism score ─────────────────────────────────
    log("Computing conservatism scores...")
    guided_rows = [r for r in rows if r["verdict"] != "NO GUIDANCE"]
    scores = []
    for r in guided_rows:
        diff_pct = r["rev_diff_pct"] or 0
        beat_score = max(0, min(33, (diff_pct + 5) / 10 * 33))

        if r["guide_low"] and r["guide_high"] and r["guide_midpoint"]:
            range_width_pct = (r["guide_high"] - r["guide_low"]) / r["guide_midpoint"] * 100
        else:
            range_width_pct = 0
        width_score = max(0, min(33, range_width_pct / 5 * 33))

        idx = guided_rows.index(r)
        recent = guided_rows[max(0, idx-3):idx+1]
        beat_rate = sum(1 for x in recent if x["verdict"] == "BEAT") / len(recent)
        consistency_score = beat_rate * 33

        total_score = round(beat_score + width_score + consistency_score)
        total_score = max(0, min(100, total_score))

        scores.append({
            "fiscal_quarter": r["fiscal_quarter"],
            "beat_score": round(beat_score, 1),
            "width_score": round(width_score, 1),
            "consistency_score": round(consistency_score, 1),
            "total_score": total_score,
            "rev_diff_pct": r["rev_diff_pct"],
        })

    for i, s in enumerate(scores):
        window = scores[max(0, i-3):i+1]
        s["rolling_4q"] = round(sum(x["total_score"] for x in window) / len(window), 1)

    if len(scores) >= 3:
        xs = list(range(len(scores)))
        ys = [s["total_score"] for s in scores]
        slope, _ = linear_regression(xs, ys)
    else:
        slope = 0

    # ── Module 3: Stock price reaction ───────────────────────────────
    log("Fetching stock prices around earnings dates...")

    def _fetch_price_reaction(r):
        """Fetch stock prices for one earnings event and compute returns."""
        fd = r["filing_date"]
        fd_dt = datetime.strptime(fd, "%Y-%m-%d")
        start = (fd_dt - timedelta(days=8)).strftime("%Y-%m-%d")
        end = (fd_dt + timedelta(days=8)).strftime("%Y-%m-%d")
        prices = fetch_stock_prices(ticker, start, end)
        if len(prices) < 3:
            return None

        prices.sort(key=lambda p: p["time"])
        dates = [p["time"][:10] for p in prices]

        fd_idx = None
        for i, d in enumerate(dates):
            if d >= fd:
                fd_idx = i
                break
        if fd_idx is None:
            return None

        pre_close = prices[max(0, fd_idx - 1)]["close"] if fd_idx > 0 else None
        post_1d = prices[min(len(prices)-1, fd_idx + 1)]["close"] if fd_idx + 1 < len(prices) else None
        post_3d = prices[min(len(prices)-1, fd_idx + 3)]["close"] if fd_idx + 3 < len(prices) else None

        ret_1d = ((post_1d / pre_close) - 1) * 100 if pre_close and post_1d else None
        ret_3d = ((post_3d / pre_close) - 1) * 100 if pre_close and post_3d else None

        return {
            "fiscal_quarter": r["fiscal_quarter"],
            "earnings_date": fd,
            "verdict": r["verdict"],
            "rev_surprise_pct": r["rev_diff_pct"],
            "pre_close": pre_close,
            "post_1d_close": post_1d,
            "ret_1d": ret_1d,
            "ret_3d": ret_3d,
        }

    price_reactions = []
    with ThreadPoolExecutor(max_workers=8) as pool:
        futures = [pool.submit(_fetch_price_reaction, r) for r in guided_rows]
        for f in as_completed(futures):
            result = f.result()
            if result:
                price_reactions.append(result)
    # Restore chronological order
    price_reactions.sort(key=lambda p: p["earnings_date"])

    valid_pairs = [(p["rev_surprise_pct"], p["ret_1d"]) for p in price_reactions
                   if p["rev_surprise_pct"] is not None and p["ret_1d"] is not None]
    corr = correlation([x[0] for x in valid_pairs], [x[1] for x in valid_pairs]) if len(valid_pairs) >= 3 else None

    # ── Module 4: Seasonal patterns ──────────────────────────────────
    log("Analyzing seasonal patterns...")
    seasonal = {}
    for r in guided_rows:
        q_num = r["fiscal_quarter"].split("-")[-1]
        if q_num not in seasonal:
            seasonal[q_num] = []
        seasonal[q_num].append(r)

    seasonal_summary = {}
    for q_num in sorted(seasonal.keys()):
        qrows = seasonal[q_num]
        diffs = [r["rev_diff_pct"] for r in qrows if r["rev_diff_pct"] is not None]
        beat_count = sum(1 for r in qrows if r["verdict"] == "BEAT")
        seasonal_summary[q_num] = {
            "n": len(qrows),
            "avg_surprise": sum(diffs) / len(diffs) if diffs else None,
            "beat_rate": beat_count / len(qrows) * 100 if qrows else None,
            "max_beat": max(diffs) if diffs else None,
            "max_miss": min(diffs) if diffs else None,
        }

    # ── Module 5: Multi-metric accuracy ──────────────────────────────
    log("Comparing accuracy across guided metrics...")
    metric_data = {"Revenue": [], "Operating Margin": []}
    for r in guided_rows:
        if r["rev_diff_pct"] is not None:
            metric_data["Revenue"].append({
                "guided": r["guide_midpoint"],
                "actual": r["actual_revenue"],
                "diff_pct": r["rev_diff_pct"],
                "verdict": r["verdict"],
            })
        if r["guide_op_margin"] is not None and r["op_margin_actual"] is not None:
            op_diff = r["op_margin_actual"] - r["guide_op_margin"]
            metric_data["Operating Margin"].append({
                "guided": r["guide_op_margin"],
                "actual": r["op_margin_actual"],
                "diff_pct": op_diff,
                "verdict": "BEAT" if op_diff > 0.5 else "MISS" if op_diff < -0.5 else "IN-RANGE",
            })

    metric_summary = {}
    for metric, entries in metric_data.items():
        if not entries:
            continue
        diffs = [e["diff_pct"] for e in entries]
        beats = sum(1 for e in entries if e["verdict"] == "BEAT")
        misses = sum(1 for e in entries if e["verdict"] == "MISS")
        metric_summary[metric] = {
            "n": len(entries),
            "avg_surprise": sum(diffs) / len(diffs),
            "avg_abs_error": sum(abs(d) for d in diffs) / len(diffs),
            "beat_rate": beats / len(entries) * 100,
            "miss_rate": misses / len(entries) * 100,
            "max_beat": max(diffs),
            "max_miss": min(diffs),
        }

    # Revenue metric name (from most recent)
    rev_metric_name = "Revenue"
    if rows:
        rev_metric_name = rows[-1].get("revenue_metric", "Revenue")

    return {
        "rows": rows,
        "guided_rows": guided_rows,
        "parsed": parsed,
        "fy_walk_rows": fy_walk_rows,
        "fy_walk": fy_walk,
        "scores": scores,
        "score_slope": slope,
        "price_reactions": price_reactions,
        "price_corr": corr,
        "seasonal_summary": seasonal_summary,
        "metric_summary": metric_summary,
        "metric_data": metric_data,
        "revenue_metric_name": rev_metric_name,
    }


# ── Excel writers ───────────────────────────────────────────────────────

def _write_header_row(ws, row, headers, widths=None):
    for c, h in enumerate(headers, 1):
        cell = ws.cell(row=row, column=c, value=h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
        cell.border = THIN_BORDER
    if widths:
        for i, w in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = w


def write_guidance_vs_actuals(wb, data, ticker):
    ws = wb.active
    ws.title = "Guidance vs Actuals"
    ws.sheet_properties.tabColor = "1F4E79"

    metric = data.get("revenue_metric_name", "Revenue")
    headers = ["Fiscal Quarter", "Earnings Date", "Guidance Given In",
               f"Guide Low ($M)", f"Guide High ($M)", f"Guide Midpoint ($M)",
               f"Actual {metric} ($M)", "Diff vs Mid ($M)", "Diff vs Mid (%)",
               "Verdict", "Gross Margin %", "GAAP Op Margin %",
               "Guided Op Margin %", f"{metric} YoY %"]
    widths = [14, 13, 16, 16, 16, 16, 18, 15, 13, 12, 13, 15, 15, 14]
    _write_header_row(ws, 1, headers, widths)

    for r_idx, row in enumerate(data["rows"], 2):
        vals = [row["fiscal_quarter"], row["filing_date"],
                row["guidance_source_q"] or "—",
                row["guide_low"], row["guide_high"], row["guide_midpoint"],
                row["actual_revenue"], row["rev_diff_vs_mid"], row["rev_diff_pct"],
                row["verdict"],
                row["gross_margin"], row["op_margin_actual"],
                row["guide_op_margin"], row.get("rev_yoy")]

        for c, v in enumerate(vals, 1):
            cell = ws.cell(row=r_idx, column=c, value=v)
            cell.border = THIN_BORDER
            cell.alignment = Alignment(horizontal="center")
            if c in (4, 5, 6, 7):
                cell.number_format = '#,##0.0'
            elif c == 8:
                cell.number_format = '+#,##0.0;-#,##0.0'
            elif c in (9, 11, 12, 13, 14):
                cell.number_format = '0.0'

        verdict = row["verdict"]
        ws.cell(row=r_idx, column=10).fill = VERDICT_FILL.get(verdict, NO_GUIDE_FILL)
        if verdict in VERDICT_FONT:
            ws.cell(row=r_idx, column=10).font = VERDICT_FONT[verdict]

    ws.freeze_panes = "D2"


def write_fy_walk(wb, data, ticker):
    ws = wb.create_sheet("FY Guidance Walk")
    ws.sheet_properties.tabColor = "2E75B6"

    headers = ["Fiscal Year", "Revision #", "Given At (Q)", "Date",
               "FY Rev Guide ($M)", "Op Margin %", "FCF Margin %",
               "Action", "Chg vs Prior ($M)", "Chg vs Prior (%)",
               "Chg vs Initial ($M)", "Chg vs Initial (%)"]
    widths = [13, 12, 15, 12, 17, 13, 13, 12, 16, 14, 16, 15]
    _write_header_row(ws, 1, headers, widths)

    action_fill = {"RAISE": RAISE_FILL, "LOWER": LOWER_FILL, "REAFFIRM": REAFFIRM_FILL, "INITIAL": EST_FILL}
    action_font = {"RAISE": Font(bold=True, color="006100"), "LOWER": Font(bold=True, color="9C0006")}

    for r_idx, row in enumerate(data["fy_walk_rows"], 2):
        vals = [row["fy_target"], row["revision_num"], row["source_q"], row["filing_date"],
                row["fy_rev"], row["fy_op_margin"], row["fy_fcf_margin"],
                row["action"], row["chg_vs_prior"], row["chg_vs_prior_pct"],
                row["chg_vs_initial"], row["chg_vs_initial_pct"]]
        for c, v in enumerate(vals, 1):
            cell = ws.cell(row=r_idx, column=c, value=v)
            cell.border = THIN_BORDER
            cell.alignment = Alignment(horizontal="center")
            if c == 5:
                cell.number_format = '#,##0.0'
            elif c in (6, 7, 10, 12):
                cell.number_format = '0.0'
            elif c in (9, 11):
                cell.number_format = '+#,##0.0;-#,##0.0'

        action = row["action"]
        action_cell = ws.cell(row=r_idx, column=8)
        if action in action_fill:
            action_cell.fill = action_fill[action]
        if action in action_font:
            action_cell.font = action_font[action]


def write_conservatism(wb, data, ticker):
    ws = wb.create_sheet("Conservatism Score")
    ws.sheet_properties.tabColor = "548235"

    headers = ["Fiscal Quarter", "Beat Magnitude (0-33)", "Range Width (0-33)",
               "Consistency (0-33)", "Total Score (0-100)", "Rolling 4Q Avg"]
    widths = [15, 20, 18, 18, 18, 15]
    _write_header_row(ws, 1, headers, widths)

    scores = data["scores"]
    for r_idx, s in enumerate(scores, 2):
        vals = [s["fiscal_quarter"], s["beat_score"], s["width_score"],
                s["consistency_score"], s["total_score"], s["rolling_4q"]]
        for c, v in enumerate(vals, 1):
            cell = ws.cell(row=r_idx, column=c, value=v)
            cell.border = THIN_BORDER
            cell.alignment = Alignment(horizontal="center")
            cell.number_format = '0.0' if c != 5 else '0'
        score_cell = ws.cell(row=r_idx, column=5)
        if s["total_score"] >= 66:
            score_cell.fill = BEAT_FILL
        elif s["total_score"] >= 50:
            score_cell.fill = RANGE_FILL
        elif s["total_score"] < 33:
            score_cell.fill = MISS_FILL


def write_market_reaction(wb, data, ticker):
    ws = wb.create_sheet("Market Reaction")
    ws.sheet_properties.tabColor = "C00000"

    headers = ["Fiscal Quarter", "Earnings Date", "Verdict", "Rev Surprise (%)",
               "Pre-Close ($)", "Post 1D Close ($)", "1-Day Return (%)", "3-Day Return (%)"]
    widths = [15, 13, 12, 16, 14, 16, 15, 15]
    _write_header_row(ws, 1, headers, widths)

    reactions = data["price_reactions"]
    for r_idx, p in enumerate(reactions, 2):
        vals = [p["fiscal_quarter"], p["earnings_date"], p["verdict"],
                p["rev_surprise_pct"], p["pre_close"], p["post_1d_close"],
                p["ret_1d"], p["ret_3d"]]
        for c, v in enumerate(vals, 1):
            cell = ws.cell(row=r_idx, column=c, value=v)
            cell.border = THIN_BORDER
            cell.alignment = Alignment(horizontal="center")
            if c in (4, 7, 8):
                cell.number_format = '+0.0;-0.0'
            elif c in (5, 6):
                cell.number_format = '#,##0.00'
        v = p["verdict"]
        if v in VERDICT_FILL:
            ws.cell(row=r_idx, column=3).fill = VERDICT_FILL[v]
        if v in VERDICT_FONT:
            ws.cell(row=r_idx, column=3).font = VERDICT_FONT[v]


def write_seasonal(wb, data, ticker):
    ws = wb.create_sheet("Seasonal Patterns")
    ws.sheet_properties.tabColor = "FFC000"

    headers = ["Quarter", "N", "Avg Surprise (%)", "Beat Rate (%)", "Max Beat (%)", "Max Miss (%)"]
    widths = [12, 8, 18, 14, 14, 14]
    _write_header_row(ws, 1, headers, widths)

    seasonal = data["seasonal_summary"]
    for r_idx, q in enumerate(sorted(seasonal.keys()), 2):
        s = seasonal[q]
        vals = [q, s["n"], s["avg_surprise"], s["beat_rate"], s["max_beat"], s["max_miss"]]
        for c, v in enumerate(vals, 1):
            cell = ws.cell(row=r_idx, column=c, value=v)
            cell.border = THIN_BORDER
            cell.alignment = Alignment(horizontal="center")
            if c in (3, 4, 5, 6):
                cell.number_format = '0.0'


def write_metric_accuracy(wb, data, ticker):
    ws = wb.create_sheet("Metric Accuracy")
    ws.sheet_properties.tabColor = "7030A0"

    headers = ["Metric", "N Quarters", "Avg Surprise", "Avg |Error|",
               "Beat Rate (%)", "Miss Rate (%)", "Max Beat", "Max Miss"]
    widths = [20, 12, 14, 14, 14, 14, 12, 12]
    _write_header_row(ws, 1, headers, widths)

    ms = data["metric_summary"]
    for r_idx, (metric, s) in enumerate(ms.items(), 2):
        unit = "pp" if "Margin" in metric else "%"
        vals = [metric, s["n"],
                f"{s['avg_surprise']:+.1f}{unit}",
                f"{s['avg_abs_error']:.1f}{unit}",
                s["beat_rate"], s["miss_rate"],
                f"{s['max_beat']:+.1f}", f"{s['max_miss']:+.1f}"]
        for c, v in enumerate(vals, 1):
            cell = ws.cell(row=r_idx, column=c, value=v)
            cell.border = THIN_BORDER
            cell.alignment = Alignment(horizontal="center")


def write_dashboard(wb, data, ticker):
    ws = wb.create_sheet("Dashboard", 0)
    ws.sheet_properties.tabColor = "1F4E79"

    guided = data["guided_rows"]
    scores = data["scores"]
    reactions = data["price_reactions"]
    fy_walk = data["fy_walk"]
    seasonal = data["seasonal_summary"]

    beats = sum(1 for r in guided if r["verdict"] == "BEAT")
    in_range = sum(1 for r in guided if r["verdict"] == "IN-RANGE")
    misses = sum(1 for r in guided if r["verdict"] == "MISS")
    total = len(guided)
    diffs = [r["rev_diff_pct"] for r in guided if r["rev_diff_pct"] is not None]
    avg_diff = sum(diffs) / len(diffs) if diffs else 0

    avg_score = sum(s["total_score"] for s in scores) / len(scores) if scores else 0
    archetype = ("Sandbagging" if avg_score > 66 else "Conservative" if avg_score > 50
                 else "Straight Shooter" if avg_score > 33 else "Aggressive")
    archetype_desc = {
        "Sandbagging": "Management systematically under-guides to create reliable beat-and-raise narratives.",
        "Conservative": "Management sets achievable targets with moderate upside built in.",
        "Straight Shooter": "Management aims for accurate guidance, resulting in a balanced beat/miss record.",
        "Aggressive": "Management sets ambitious targets, frequently requiring downward revisions.",
    }

    ws.merge_cells("A1:F1")
    ws.cell(row=1, column=1, value=f"{ticker} — Management Guidance Philosophy Profile").font = Font(bold=True, size=16, color="1F4E79")

    r = 3
    ws.merge_cells(f"A{r}:F{r}")
    ws.cell(row=r, column=1, value="GUIDANCE PHILOSOPHY CLASSIFICATION").font = SECTION_FONT
    ws.cell(row=r, column=1).fill = SECTION_FILL
    r += 1
    ws.cell(row=r, column=1, value="Archetype:").font = LABEL_FONT
    ws.cell(row=r, column=2, value=archetype).font = Font(bold=True, size=14, color="1F4E79")
    r += 1
    ws.merge_cells(f"A{r}:F{r}")
    ws.cell(row=r, column=1, value=archetype_desc.get(archetype, "")).font = Font(italic=True, size=10)

    r += 2
    ws.merge_cells(f"A{r}:F{r}")
    ws.cell(row=r, column=1, value="KEY METRICS").font = SECTION_FONT
    ws.cell(row=r, column=1).fill = SECTION_FILL
    r += 1

    metrics = [
        ("Quarters Analyzed", str(total)),
        ("Beat / In-Range / Miss", f"{beats} / {in_range} / {misses}"),
        ("Beat Rate", f"{beats/total*100:.0f}%" if total else "N/A"),
        ("Avg Surprise vs Midpoint", f"{avg_diff:+.1f}%"),
        ("Conservatism Score (avg)", f"{avg_score:.0f}/100"),
        ("Score Trend", f"{data['score_slope']:+.1f} pts/quarter"),
    ]

    total_raises = []
    for fy, revs in fy_walk.items():
        if len(revs) >= 2:
            initial = revs[0]["fy_rev"]
            final = revs[-1]["fy_rev"]
            total_raises.append((final / initial - 1) * 100)
    if total_raises:
        avg_fy_raise = sum(total_raises) / len(total_raises)
        metrics.append(("Avg FY Guide Total Raise", f"{avg_fy_raise:+.1f}%"))

    beat_rets = [p["ret_1d"] for p in reactions if p["verdict"] == "BEAT" and p["ret_1d"] is not None]
    if beat_rets:
        metrics.append(("Avg 1D Return on Beat", f"{sum(beat_rets)/len(beat_rets):+.1f}%"))

    if data["price_corr"] is not None:
        metrics.append(("Surprise-Return Correlation", f"r = {data['price_corr']:.2f}"))

    if seasonal:
        valid_seasonal = {k: v for k, v in seasonal.items() if v["avg_surprise"] is not None}
        if valid_seasonal:
            best_q = max(valid_seasonal.items(), key=lambda x: x[1]["avg_surprise"])
            metrics.append(("Most Conservative Quarter", f"{best_q[0]} (+{best_q[1]['avg_surprise']:.1f}% avg)"))

    for label, val in metrics:
        ws.cell(row=r, column=1, value=label).font = LABEL_FONT
        ws.cell(row=r, column=3, value=val)
        r += 1

    ws.column_dimensions["A"].width = 26
    ws.column_dimensions["B"].width = 10
    ws.column_dimensions["C"].width = 30
    ws.column_dimensions["D"].width = 14
    ws.column_dimensions["E"].width = 14
    ws.column_dimensions["F"].width = 14


def write_excel(data: dict, ticker: str) -> str:
    wb = Workbook()
    write_guidance_vs_actuals(wb, data, ticker)
    write_fy_walk(wb, data, ticker)
    write_conservatism(wb, data, ticker)
    write_market_reaction(wb, data, ticker)
    write_seasonal(wb, data, ticker)
    write_metric_accuracy(wb, data, ticker)
    write_dashboard(wb, data, ticker)

    filename = f"{ticker}_guidance_analysis.xlsx"
    filepath = f"/Users/yimengwang/Desktop/guides_rep/{filename}"
    wb.save(filepath)
    return filepath


def write_excel_to_bytes(data: dict, ticker: str) -> bytes:
    """Write Excel to bytes buffer (for Streamlit download)."""
    from io import BytesIO
    wb = Workbook()
    write_guidance_vs_actuals(wb, data, ticker)
    write_fy_walk(wb, data, ticker)
    write_conservatism(wb, data, ticker)
    write_market_reaction(wb, data, ticker)
    write_seasonal(wb, data, ticker)
    write_metric_accuracy(wb, data, ticker)
    write_dashboard(wb, data, ticker)
    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ── Main ────────────────────────────────────────────────────────────────

def main():
    ticker = sys.argv[1].upper() if len(sys.argv) > 1 else "SNOW"
    print(f"\n{'='*60}")
    print(f"  Guidance Philosophy Analyzer — {ticker}")
    print(f"{'='*60}\n")

    data = build_all_data(ticker)

    guided = data["guided_rows"]
    beats = sum(1 for r in guided if r["verdict"] == "BEAT")
    total = len(guided)

    print(f"\n  {total} quarters with management guidance data")
    if total:
        print(f"  Beat rate: {beats}/{total} ({beats/total*100:.0f}%)")
        if data["scores"]:
            avg = sum(s["total_score"] for s in data["scores"]) / len(data["scores"])
            print(f"  Conservatism score: {avg:.0f}/100")

    filepath = write_excel(data, ticker)
    print(f"\n  Spreadsheet saved: {filepath}\n")


if __name__ == "__main__":
    main()
